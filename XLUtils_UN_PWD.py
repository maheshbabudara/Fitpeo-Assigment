import openpyxl
from openpyxl.styles import PatternFill

file=("C:\Python_Selenium\Selenium\PythonSelenium\WiteData_practise.xlsx")
def getrow(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row
def getcol(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column
def readata(file,sheetname,row,col):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.cell(row,col).value)
def writedata(file,sheetname,row,col,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row,col).value=data
    workbook.save(file)

def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenfil = PatternFill(start_color='60b212', end_color='60b212', fill_type="solid")
    sheet.cell(rownum, columnno).fill = greenfil
    workbook.save(file)


def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redfil = PatternFill(start_color='ff0000', end_color='ff0000', fill_type="solid")
    sheet.cell(rownum, columnno).fill = redfil
    workbook.save(file)
