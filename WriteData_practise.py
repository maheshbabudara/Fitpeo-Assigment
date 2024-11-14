import openpyxl

file_path=("C:\Python_Selenium\Selenium\PythonSelenium\kosar.xlsx")

workbook=openpyxl.load_workbook(file_path)
sheet=workbook["da"]


#Write various data in excel sheet:
sheet.cell(1,1).value="Name"
sheet.cell(1,2).value='Age'
sheet.cell(1,3).value='city'


sheet.cell(2,1).value="Rony"
sheet.cell(2,2).value=23
sheet.cell(2,3).value='Nandyala'


sheet.cell(3,1).value="Mahesh"
sheet.cell(3,2).value=24
sheet.cell(3,3).value='Nandy'

sheet.cell(4,1).value="Kolar"
sheet.cell(4,2).value=25
sheet.cell(4,3).value='kimi'

workbook.save(file_path)


#Read Data from Excel sheet:
file_path=("C:\Python_Selenium\Selenium\PythonSelenium\kosar.xlsx")
workbook=openpyxl.load_workbook(file_path)
sheet=workbook['da']

rows=sheet.max_row
column=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value, end=' ')
    print()


import openpyxl

from openpyxl.styles import PatternFill

file_path= ("C:\Python_Selenium\Selenium\PythonSelenium\WiteData_practise.xlsx")

workbook= openpyxl.load_workbook(file_path)
sheet=workbook.active

#Heading Row:
sheet.cell(1,1).value="Name"
sheet.cell(1,2).value="Password"

#Values:
sheet.cell(2,1).value="Admin"
sheet.cell(2,2).value="admin123"

sheet.cell(3,1).value="mahesh"
sheet.cell(3,2).value="123"

sheet.cell(4,1).value="rony"
sheet.cell(4,2).value="12345"

#Saving workbhook:
workbook.save(file_path)

#Read data of Excel sheet:
file=("C:\Python_Selenium\Selenium\PythonSelenium\WiteData_practise.xlsx")
workbook=openpyxl.load_workbook(file)
sheet=workbook.active
rows= sheet.max_row
col=sheet.max_column
for r in range(1,rows+1):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value, end=" ")
    print()


