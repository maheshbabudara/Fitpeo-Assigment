import openpyxl

file_path=("C:\Python_Selenium\Selenium\PythonSelenium\Testdata1.xlsx")
workbook=openpyxl.load_workbook(file_path)
sheet=workbook.active

# sheet.cell(r,c).value
sheet.cell(1,1).value="rony"
sheet.cell(1,2).value=23
sheet.cell(1,3).value="Chinka"

sheet.cell(2,1).value="rony1"
sheet.cell(2,2).value=22
sheet.cell(2,3).value="Cha"

sheet.cell(3,1).value="ro"
sheet.cell(3,2).value=28
sheet.cell(3,3).value="Chee"

workbook.save(file_path)