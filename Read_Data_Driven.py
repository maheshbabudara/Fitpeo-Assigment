import openpyxl

#structure of excel: File-->workbook-->sheet-->Rows-->cell
#file_path:
file_path=("C:\Python_Selenium\Selenium\PythonSelenium\Testdata.xlsx")

#workbook:
workbook=openpyxl.load_workbook(file_path)

#sheet:
sheet=workbook["Details"]

#rows:
sheet_rows=sheet.max_row
sheet_col=sheet.max_column

#Reading data of all rows and Columns:
for r in range(1,sheet_rows+1):
    for c in range(1,sheet_col+1):
        print(sheet.cell(r,c).value, end='  ')
    print()