import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum, columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)

def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenfil = PatternFill(start_color='60b212', end_color='60b212', fill_type="solid")
    sheet.cell(rownum, columnno).fill = greenfil
    workbook.save(file)

def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redfil = PatternFill(start_color='ff0000', end_color='ff0000', fill_type="solid")
    sheet.cell(rownum, columnno).fill = redfil
    workbook.save(file)
