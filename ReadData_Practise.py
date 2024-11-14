from selenium import webdriver
import openpyxl

from selenium.webdriver.common.by import By
from time import sleep

def setup():
    from selenium.webdriver.chrome.service import Service
    serv_obj=Service("C:\Python_Selenium\Drivers\chromedriver-win64\chromedriver.exe")
    driver=webdriver.Chrome(service=serv_obj)
    return driver
my_driver=setup()

#File:
file_path=("C:\Python_Selenium\Selenium\PythonSelenium\Testdata.xlsx")

#Workbook:
workbook=openpyxl.load_workbook(file_path)

#sheet:
sheet=workbook["Details"]

#rows and columns:
rows=sheet.max_row
column=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value, end=' ')
    print()



