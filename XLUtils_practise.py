# import openpyxl as p
# from openpyxl.styles import PatternFill
#
# def getrowcount(file,sheetName):
#     workbook= p.load_workbook(file)
#     sheet=workbook[sheetName]
#     return sheet.max_row
#
# def getcolcount(file,sheetName):
#     workbook=p.load_workbook(file)
#     sheet=workbook[sheetName]
#     return sheet.max_column
#
# def readdata(file,sheetName,rowno,colno):
#     workbook=p.load_workbook(file)
#     sheet=workbook[sheetName]
#     return sheet.cell(rowno,colno).value
#
# def writedata(file,sheetName,rowno,colno,data):
#     workbook=p.load_workbook(file)
#     sheet=workbook[sheetName]
#     sheet.cell(rowno,colno).value= data
#     workbook.save(file)
#
# def greencolor(file,sheetName,rowno,colno):
#     workbook=p.load_workbook(file)
#     sheet=workbook[sheetName]
#     greenfil= PatternFill(start_color="60b212", end_color="60b212", fill_type="solid")
#     sheet.cell(rowno,colno).fill=greenfil
#     workbook.save(file)
#
# def redcolor(file,sheetName,rowno,colno):
#     workbook=p.load_workbook(file)
#     sheet=workbook[sheetName]
#     redfil=PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
#     sheet.cell(rowno,colno).fill=redfil
#     workbook.save(file)

import openpyxl
from openpyxl.styles import PatternFill

def rowcount(sheetname,file):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_row
def colcount(sheetname,file):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.max_column
def writedata(sheetname,file,row,col,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row,col).value=data
    workbook.save(file)

def readdata(sheetname,file,row,col):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(row,col).value
def greenfill(sheetname,file,row,col):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    green=PatternFill(start_color="60b212", end_color="60b212", fill_type="solid")
    sheet.cell(row,col).fill=green
    workbook.save(file)
def redfill(sheetname,file,row,col):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    red=PatternFill(start_color="ff0000",end_color="ff0000", fill_type="solid")
    sheet.cell(row,col).fill=red
    workbook.save(file)

